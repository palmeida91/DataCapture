#!/usr/bin/env python3
"""
Get Shift Downtime Data
Queries technical_availability table for fault/blocked/starved time per station.
Uses MAX per hour then SUM across hours to get true shift totals.

Usage:
    # Single day, single shift
    python3.12 get_shift_downtime.py --date 2026-02-18 --shift 2

    # All active sequences, single day, single shift
    python3.12 get_shift_downtime.py --date 2026-02-18 --shift 2

    # Date range, specific shift
    python3.12 get_shift_downtime.py -s 50 51 --from 2026-02-17 --to 2026-02-21 --shift 1

    # All shifts in a day
    python3.12 get_shift_downtime.py --date 2026-02-18

    # Export to CSV
    python3.12 get_shift_downtime.py --date 2026-02-18 --shift 2 --format csv

Dependencies:
    pip3.12 install psycopg2-binary --break-system-packages
    pip3.12 install openpyxl --break-system-packages  (only for xlsx output)
"""

import argparse
import csv
import json
import os
import sys
from datetime import datetime, date, timedelta, time as dt_time
from typing import Dict, List, Optional, Tuple

import psycopg2
import psycopg2.extras


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(SCRIPT_DIR, "config", "collector_config.json")
DEFAULT_OUTPUT_DIR = os.path.join(SCRIPT_DIR, "reports")

# Shift definitions
SHIFTS = {
    'monday_to_thursday': {
        1: {'start': '06:00', 'end': '14:00'},
        2: {'start': '14:00', 'end': '22:00'},
        3: {'start': '22:00', 'end': '06:00'},
    },
    'friday': {
        1: {'start': '06:00', 'end': '13:30'},
        2: {'start': '13:30', 'end': '21:00'},
        3: {'start': '21:00', 'end': '04:30'},
    }
}


def load_db_config() -> dict:
    """Load database credentials from collector_config.json."""
    if not os.path.exists(CONFIG_PATH):
        print(f"ERROR: Config file not found: {CONFIG_PATH}")
        print("Make sure you're running from the DataCapture/ directory.")
        sys.exit(1)

    with open(CONFIG_PATH, "r") as f:
        config = json.load(f)

    db = config.get("database", {})
    return {
        "host": db.get("host", "localhost"),
        "port": db.get("port", 5432),
        "dbname": db.get("database", "production"),
        "user": db.get("user", "collector"),
        "password": db.get("password", ""),
    }


def get_shift_boundaries(shift_number: int, day_date: date) -> Tuple[datetime, datetime]:
    """
    Get shift start/end as full datetime objects.
    Handles overnight shifts (shift 3) by adding a day to the end time.
    """
    is_friday = day_date.isoweekday() == 5
    shift_key = 'friday' if is_friday else 'monday_to_thursday'
    shift_def = SHIFTS[shift_key][shift_number]

    start_h, start_m = map(int, shift_def['start'].split(':'))
    end_h, end_m = map(int, shift_def['end'].split(':'))

    shift_start = datetime.combine(day_date, dt_time(start_h, start_m))
    shift_end = datetime.combine(day_date, dt_time(end_h, end_m))

    # Overnight shift: end is next day
    if shift_end <= shift_start:
        shift_end += timedelta(days=1)

    return shift_start, shift_end


# ---------------------------------------------------------------------------
# Database query
# ---------------------------------------------------------------------------

def fetch_shift_downtime(
    date_from: date,
    date_to: date,
    shift_numbers: List[int],
    sequence_ids: Optional[List[int]] = None,
) -> List[dict]:
    """
    Query downtime per station per shift.
    Uses MAX per hour (since PLC updates hourly) then SUM across hours
    to get true shift totals for fault/blocked/starved.
    """
    db_config = load_db_config()

    try:
        conn = psycopg2.connect(**db_config)
        conn.autocommit = True
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    except psycopg2.OperationalError as e:
        print(f"ERROR: Database connection failed:\n  {e}")
        sys.exit(1)

    all_results = []

    # Iterate each day in range
    current_date = date_from
    while current_date <= date_to:
        # Skip weekends
        if current_date.isoweekday() > 5:
            current_date += timedelta(days=1)
            continue

        for shift_num in shift_numbers:
            shift_start, shift_end = get_shift_boundaries(shift_num, current_date)

            # Build query with simple datetime range
            seq_filter = ""
            query_params = [shift_start, shift_end]

            if sequence_ids:
                seq_filter = "AND ta.sequence_id = ANY(%s)"
                query_params.append(sequence_ids)

            query = f"""
                WITH hourly_max AS (
                    SELECT
                        ta.sequence_id,
                        date_trunc('hour', ta.time) AS hour_bucket,
                        MAX(ta.fault_time_seconds) AS fault_sec,
                        MAX(ta.blocked_time_seconds) AS blocked_sec,
                        MAX(ta.starved_time_seconds) AS starved_sec
                    FROM technical_availability ta
                    JOIN sequences s ON ta.sequence_id = s.sequence_id
                    WHERE s.is_active = true
                        AND ta.time >= %s
                        AND ta.time < %s
                        {seq_filter}
                    GROUP BY ta.sequence_id, date_trunc('hour', ta.time)
                )
                SELECT
                    s.sequence_name,
                    s.sequence_id,
                    ROUND(COALESCE(SUM(hm.fault_sec), 0)::numeric / 60, 1) AS fault_min,
                    ROUND(COALESCE(SUM(hm.blocked_sec), 0)::numeric / 60, 1) AS blocked_min,
                    ROUND(COALESCE(SUM(hm.starved_sec), 0)::numeric / 60, 1) AS starved_min,
                    ROUND(COALESCE(SUM(hm.fault_sec + hm.blocked_sec + hm.starved_sec), 0)::numeric / 60, 1)
                        AS total_downtime_min
                FROM hourly_max hm
                JOIN sequences s ON hm.sequence_id = s.sequence_id
                GROUP BY s.sequence_name, s.sequence_id
                ORDER BY total_downtime_min DESC;
            """

            cur.execute(query, query_params)
            rows = cur.fetchall()

            for row in rows:
                result = dict(row)
                result['shift_number'] = shift_num
                result['shift_date'] = current_date.isoformat()
                all_results.append(result)

        current_date += timedelta(days=1)

    cur.close()
    conn.close()

    return all_results


# ---------------------------------------------------------------------------
# Output formatting
# ---------------------------------------------------------------------------

def print_results(results: List[dict]):
    """Print results as a formatted table to terminal."""
    if not results:
        print("No data found.")
        return

    # Header
    print(f"\n{'Date':<12} {'Shift':<6} {'Station':<25} {'Fault (min)':<12} "
          f"{'Blocked (min)':<14} {'Starved (min)':<14} {'Total (min)':<12}")
    print("-" * 95)

    for r in results:
        print(f"{r['shift_date']:<12} {r['shift_number']:<6} {r['sequence_name']:<25} "
              f"{r['fault_min']:<12} {r['blocked_min']:<14} {r['starved_min']:<14} "
              f"{r['total_downtime_min']:<12}")

    # Totals
    total_fault = sum(float(r['fault_min']) for r in results)
    total_blocked = sum(float(r['blocked_min']) for r in results)
    total_starved = sum(float(r['starved_min']) for r in results)
    total_all = sum(float(r['total_downtime_min']) for r in results)

    print("-" * 95)
    print(f"{'TOTAL':<12} {'':<6} {'':<25} {total_fault:<12.1f} "
          f"{total_blocked:<14.1f} {total_starved:<14.1f} {total_all:<12.1f}")
    print()


def export_csv(results: List[dict], output_path: str):
    """Export results to CSV."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    headers = ['shift_date', 'shift_number', 'sequence_id', 'sequence_name',
               'fault_min', 'blocked_min', 'starved_min', 'total_downtime_min']

    with open(output_path, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(results)

    print(f"✔ Exported {len(results)} rows to {output_path}")


def export_xlsx(results: List[dict], output_path: str):
    """Export results to Excel with formatting."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl is required for Excel export.")
        print("Install it: pip3.12 install openpyxl --break-system-packages")
        sys.exit(1)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Shift Downtime"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ['Date', 'Shift', 'Seq ID', 'Station', 'Fault (min)',
               'Blocked (min)', 'Starved (min)', 'Total Downtime (min)']
    keys = ['shift_date', 'shift_number', 'sequence_id', 'sequence_name',
            'fault_min', 'blocked_min', 'starved_min', 'total_downtime_min']

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row_idx, r in enumerate(results, start=2):
        for col_idx, key in enumerate(keys, start=1):
            val = r.get(key, '')
            if key in ('fault_min', 'blocked_min', 'starved_min', 'total_downtime_min'):
                val = float(val)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx >= 5:
                cell.alignment = Alignment(horizontal="center")

    # Auto-size columns
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(headers[col_idx - 1])
        for row_idx in range(2, min(102, 2 + len(results))):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = max_len + 4

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"✔ Exported {len(results)} rows to {output_path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Get shift downtime data (fault/blocked/starved) per station.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3.12 get_shift_downtime.py --date 2026-02-18 --shift 2
  python3.12 get_shift_downtime.py -s 50 51 --date 2026-02-18 --shift 2
  python3.12 get_shift_downtime.py --from 2026-02-17 --to 2026-02-21 --shift 1
  python3.12 get_shift_downtime.py --date 2026-02-18 --format csv
  python3.12 get_shift_downtime.py -s 50 --from 2026-02-17 --to 2026-02-21 --shift 1 2 --format xlsx
        """,
    )
    parser.add_argument(
        "--sequences", "-s",
        nargs="+",
        type=int,
        default=None,
        help="Sequence IDs to query (default: all active sequences)",
    )
    parser.add_argument(
        "--date", "-d",
        type=str,
        default=None,
        help="Single date (YYYY-MM-DD)",
    )
    parser.add_argument(
        "--from",
        dest="date_from",
        type=str,
        default=None,
        help="Start date (YYYY-MM-DD)",
    )
    parser.add_argument(
        "--to",
        dest="date_to",
        type=str,
        default=None,
        help="End date (YYYY-MM-DD)",
    )
    parser.add_argument(
        "--shift",
        nargs="+",
        type=int,
        default=None,
        help="Shift number(s): 1, 2, 3 (default: all shifts)",
    )
    parser.add_argument(
        "--format", "-f",
        type=str,
        default=None,
        choices=["csv", "xlsx"],
        help="Export format (default: print to terminal)",
    )
    parser.add_argument(
        "--output", "-o",
        type=str,
        default=None,
        help="Output file path (default: reports/shift_downtime_...)",
    )

    args = parser.parse_args()

    # --- Resolve date range ---
    if args.date:
        date_from = date.fromisoformat(args.date)
        date_to = date_from
    elif args.date_from and args.date_to:
        date_from = date.fromisoformat(args.date_from)
        date_to = date.fromisoformat(args.date_to)
    elif args.date_from or args.date_to:
        print("ERROR: --from and --to must be used together.")
        sys.exit(1)
    else:
        print("ERROR: Provide either --date or --from/--to.")
        sys.exit(1)

    # --- Resolve shifts ---
    shift_numbers = args.shift if args.shift else [1, 2, 3]
    for s in shift_numbers:
        if s not in (1, 2, 3):
            print(f"ERROR: Invalid shift number {s}. Must be 1, 2, or 3.")
            sys.exit(1)

    # --- Run query ---
    print(f"Querying downtime data...")
    print(f"  Date range : {date_from} to {date_to}")
    print(f"  Shifts     : {shift_numbers}")
    print(f"  Sequences  : {args.sequences or 'all active'}")
    print()

    results = fetch_shift_downtime(date_from, date_to, shift_numbers, args.sequences)

    if not results:
        print("No data found for the given parameters.")
        return

    # --- Output ---
    if args.format:
        # Resolve output path
        if args.output:
            output_path = args.output
            if not os.path.dirname(output_path):
                output_path = os.path.join(DEFAULT_OUTPUT_DIR, output_path)
        else:
            seq_str = "_".join(str(s) for s in args.sequences) if args.sequences else "all"
            shift_str = "_".join(str(s) for s in shift_numbers)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            ext = args.format
            filename = f"shift_downtime_seq{seq_str}_shift{shift_str}_{timestamp}.{ext}"
            output_path = os.path.join(DEFAULT_OUTPUT_DIR, filename)

        if args.format == "csv":
            export_csv(results, output_path)
        else:
            export_xlsx(results, output_path)
    else:
        # Print to terminal
        print_results(results)


if __name__ == "__main__":
    main()
