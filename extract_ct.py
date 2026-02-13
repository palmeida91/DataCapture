#!/usr/bin/env python3
"""
Extract Cycle Time Data to CSV or Excel
Queries cycle_times table from PostgreSQL and exports to CSV or multi-tab XLSX.

Usage:
    python3.12 extract_ct.py --sequences 50 51 --time 8h --format csv
    python3.12 extract_ct.py --sequences 50 51 --time 8h --format xlsx
    python3.12 extract_ct.py --sequences 50 51 --from "2026-02-10 06:00" --to "2026-02-10 14:00" --format xlsx
    python3.12 extract_ct.py --sequences 50 51 --time 24h --format csv --output my_data.csv

Dependencies:
    pip3.12 install psycopg2-binary openpyxl --break-system-packages
"""

import argparse
import csv
import json
import os
import re
import sys
from datetime import datetime, timedelta

import psycopg2


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(SCRIPT_DIR, "config", "collector_config.json")
DEFAULT_OUTPUT_DIR = os.path.join(SCRIPT_DIR, "reports")

HEADERS = [
    "time",
    "sequence_id",
    "sequence_name",
    "cycle_time_seconds",
    "desired_cycle_time_seconds",
    "deviation_seconds",
    "deviation_percent",
]


def load_db_config() -> dict:
    """Load database credentials from collector_config.json."""
    if not os.path.exists(CONFIG_PATH):
        print(f"ERROR: Config file not found: {CONFIG_PATH}")
        print("Make sure you're running from the DataCapture/ directory.")
        sys.exit(1)

    with open(CONFIG_PATH, "r") as f:
        config = json.load(f)

    db = config.get("database", {})
    return {
        "host": db.get("host", "localhost"),
        "port": db.get("port", 5432),
        "database": db.get("database", "production"),
        "user": db.get("user", "collector"),
        "password": db.get("password", ""),
    }


# ---------------------------------------------------------------------------
# Time parsing
# ---------------------------------------------------------------------------

TIME_PATTERN = re.compile(r"^(\d+)\s*(m|h|d)$", re.IGNORECASE)


def parse_lookback(value: str) -> timedelta:
    """Parse a lookback string like '8h', '30m', '2d' into a timedelta."""
    match = TIME_PATTERN.match(value.strip())
    if not match:
        print(f"ERROR: Invalid time format '{value}'. Use e.g. 30m, 8h, 2d")
        sys.exit(1)

    amount = int(match.group(1))
    unit = match.group(2).lower()

    if unit == "m":
        return timedelta(minutes=amount)
    elif unit == "h":
        return timedelta(hours=amount)
    elif unit == "d":
        return timedelta(days=amount)


def parse_datetime(value: str) -> datetime:
    """Parse a datetime string in common formats."""
    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue

    print(f"ERROR: Cannot parse datetime '{value}'.")
    print("Accepted formats: '2026-02-10 06:00:00', '2026-02-10 06:00', '2026-02-10'")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Database query
# ---------------------------------------------------------------------------

def fetch_cycle_times(
    sequences: list[int],
    time_from: datetime,
    time_to: datetime,
) -> list[tuple]:
    """Query cycle_times with duplicate filtering. Returns list of row tuples."""

    db_config = load_db_config()

    query = """
        WITH ranked AS (
            SELECT
                ct.time,
                ct.sequence_id,
                s.sequence_name,
                ct.cycle_time_seconds,
                ct.desired_cycle_time_seconds,
                ct.deviation_seconds,
                ct.deviation_percent,
                LAG(ct.cycle_time_seconds) OVER (
                    PARTITION BY ct.sequence_id ORDER BY ct.time
                ) AS prev_cycle_time
            FROM cycle_times ct
            LEFT JOIN sequences s ON ct.sequence_id = s.sequence_id
            WHERE ct.sequence_id = ANY(%s)
              AND ct.time >= %s
              AND ct.time <= %s
        )
        SELECT
            time,
            sequence_id,
            sequence_name,
            cycle_time_seconds,
            desired_cycle_time_seconds,
            deviation_seconds,
            deviation_percent
        FROM ranked
        WHERE prev_cycle_time IS NULL
           OR cycle_time_seconds != prev_cycle_time
        ORDER BY time ASC, sequence_id ASC;
    """

    try:
        conn = psycopg2.connect(**db_config)
        cur = conn.cursor()
        cur.execute(query, (sequences, time_from, time_to))
        rows = cur.fetchall()
    except psycopg2.OperationalError as e:
        print(f"ERROR: Database connection failed:\n  {e}")
        sys.exit(1)
    finally:
        if "cur" in locals():
            cur.close()
        if "conn" in locals():
            conn.close()

    return rows


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------

def export_csv(rows: list[tuple], output_path: str) -> int:
    """Write rows to a CSV file. Returns row count."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    with open(output_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        for row in rows:
            writer.writerow(row)

    return len(rows)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_xlsx(rows: list[tuple], output_path: str) -> int:
    """Write rows to a multi-tab Excel workbook with summary chart. Returns row count."""
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl is required for Excel export.")
        print("Install it: pip3.12 install openpyxl --break-system-packages")
        sys.exit(1)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # -- Group rows by sequence --
    seq_data = {}  # {sequence_id: {"name": str, "rows": [tuples]}}
    for row in rows:
        seq_id = row[1]
        seq_name = row[2] or f"Sequence {seq_id}"
        if seq_id not in seq_data:
            seq_data[seq_id] = {"name": seq_name, "rows": []}
        seq_data[seq_id]["rows"].append(row)

    # -- Styles --
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    summary_title_font = Font(bold=True, size=14)
    summary_header_font = Font(bold=True, color="FFFFFF", size=10)
    summary_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    wb = Workbook()

    # =====================================================================
    # Summary tab
    # =====================================================================
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Title
    ws_summary["A1"] = "Cycle Time Summary"
    ws_summary["A1"].font = summary_title_font
    ws_summary.merge_cells("A1:H1")

    # Summary table headers
    summary_headers = [
        "Sequence ID", "Sequence Name", "Cycles",
        "Avg (s)", "Median (s)", "Min (s)", "Max (s)", "Target (s)",
    ]
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws_summary.cell(row=3, column=col_idx, value=header)
        cell.font = summary_header_font
        cell.fill = summary_header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Summary data rows (sorted by sequence_id)
    summary_rows = []
    for seq_id in sorted(seq_data.keys()):
        info = seq_data[seq_id]
        cycle_times = [r[3] for r in info["rows"] if r[3] is not None]
        if not cycle_times:
            continue

        sorted_ct = sorted(cycle_times)
        n = len(sorted_ct)
        median = sorted_ct[n // 2] if n % 2 == 1 else (sorted_ct[n // 2 - 1] + sorted_ct[n // 2]) / 2
        target = info["rows"][0][4]  # desired_cycle_time_seconds

        summary_rows.append({
            "seq_id": seq_id,
            "name": info["name"],
            "count": n,
            "avg": round(sum(cycle_times) / n, 2),
            "median": round(median, 2),
            "min": round(min(cycle_times), 2),
            "max": round(max(cycle_times), 2),
            "target": target,
        })

    for row_idx, sr in enumerate(summary_rows, start=4):
        values = [
            sr["seq_id"], sr["name"], sr["count"],
            sr["avg"], sr["median"], sr["min"], sr["max"], sr["target"],
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx >= 3:
                cell.alignment = Alignment(horizontal="center")

    # Auto-size summary columns
    for col_idx in range(1, len(summary_headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(summary_headers[col_idx - 1]))
        for row_idx in range(4, 4 + len(summary_rows)):
            cell_val = ws_summary.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        ws_summary.column_dimensions[col_letter].width = max_len + 4

    # -- Bar chart: average cycle time per sequence --
    if summary_rows:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Average Cycle Time per Sequence"
        chart.y_axis.title = "Cycle Time (s)"
        chart.x_axis.title = "Sequence"

        data_start_row = 4
        data_end_row = 3 + len(summary_rows)

        # Average values (column 4)
        data_ref = Reference(ws_summary, min_col=4, min_row=3, max_row=data_end_row)
        cats_ref = Reference(ws_summary, min_col=2, min_row=data_start_row, max_row=data_end_row)

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        chart.width = 22
        chart.height = 14

        # Target line as second series (column 8)
        target_ref = Reference(ws_summary, min_col=8, min_row=3, max_row=data_end_row)
        chart.add_data(target_ref, titles_from_data=True)

        # Style: bars blue, target red
        chart.series[0].graphicalProperties.solidFill = "4472C4"
        chart.series[1].graphicalProperties.solidFill = "FF0000"

        chart_row = data_end_row + 2
        ws_summary.add_chart(chart, f"A{chart_row}")

    # =====================================================================
    # Per-sequence tabs
    # =====================================================================
    for seq_id in sorted(seq_data.keys()):
        info = seq_data[seq_id]
        # Sheet name max 31 chars in Excel
        sheet_name = f"Seq {seq_id} - {info['name']}"[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Headers
        for col_idx, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        for row_idx, row in enumerate(info["rows"], start=2):
            for col_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border

        # Auto-size columns
        for col_idx in range(1, len(HEADERS) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(HEADERS[col_idx - 1])
            # Sample first 100 rows for width
            for row_idx in range(2, min(102, 2 + len(info["rows"]))):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val is not None:
                    max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[col_letter].width = max_len + 3

        # Freeze header row
        ws.freeze_panes = "A2"

    wb.save(output_path)
    return len(rows)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def build_output_filename(sequences: list[int], time_label: str, fmt: str) -> str:
    """Generate a default output filename."""
    seq_str = "_".join(str(s) for s in sequences)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    ext = "xlsx" if fmt == "xlsx" else "csv"
    return f"cycle_times_seq{seq_str}_{time_label}_{timestamp}.{ext}"


def main():
    parser = argparse.ArgumentParser(
        description="Extract cycle time data from PostgreSQL to CSV or Excel.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3.12 extract_ct.py -s 50 51 --time 8h --format csv
  python3.12 extract_ct.py -s 50 51 --time 8h --format xlsx
  python3.12 extract_ct.py -s 50 51 --from "2026-02-10 06:00" --to "2026-02-10 14:00" --format xlsx
  python3.12 extract_ct.py -s 50 --time 2d --format csv --output shift_analysis.csv
        """,
    )
    parser.add_argument(
        "--sequences", "-s",
        nargs="+",
        type=int,
        required=True,
        help="Sequence IDs to extract (e.g. 50 51 52)",
    )
    parser.add_argument(
        "--time", "-t",
        type=str,
        default=None,
        help="Lookback period from now (e.g. 30m, 8h, 2d)",
    )
    parser.add_argument(
        "--from",
        dest="time_from",
        type=str,
        default=None,
        help="Start datetime (e.g. '2026-02-10 06:00')",
    )
    parser.add_argument(
        "--to",
        dest="time_to",
        type=str,
        default=None,
        help="End datetime (e.g. '2026-02-10 14:00')",
    )
    parser.add_argument(
        "--format", "-f",
        type=str,
        required=True,
        choices=["csv", "xlsx"],
        help="Output format: csv (flat file) or xlsx (multi-tab with summary chart)",
    )
    parser.add_argument(
        "--output", "-o",
        type=str,
        default=None,
        help="Output file path (default: reports/cycle_times_seq...)",
    )

    args = parser.parse_args()

    # --- Resolve time range ---
    now = datetime.now()

    if args.time and (args.time_from or args.time_to):
        print("ERROR: Use either --time OR --from/--to, not both.")
        sys.exit(1)

    if args.time:
        delta = parse_lookback(args.time)
        time_from = now - delta
        time_to = now
        time_label = args.time.strip()
    elif args.time_from and args.time_to:
        time_from = parse_datetime(args.time_from)
        time_to = parse_datetime(args.time_to)
        time_label = "custom"
    elif args.time_from or args.time_to:
        print("ERROR: --from and --to must be used together.")
        sys.exit(1)
    else:
        print("ERROR: Provide either --time or --from/--to.")
        sys.exit(1)

    # --- Resolve output path ---
    if args.output:
        output_path = args.output
        # If just a filename (no directory), put it in reports/
        if not os.path.dirname(output_path):
            output_path = os.path.join(DEFAULT_OUTPUT_DIR, output_path)
    else:
        filename = build_output_filename(args.sequences, time_label, args.format)
        output_path = os.path.join(DEFAULT_OUTPUT_DIR, filename)

    # --- Fetch data ---
    print(f"Extracting cycle times...")
    print(f"  Sequences : {args.sequences}")
    print(f"  From      : {time_from.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  To        : {time_to.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Format    : {args.format}")
    print(f"  Output    : {output_path}")
    print()

    rows = fetch_cycle_times(args.sequences, time_from, time_to)

    if not rows:
        print("No data found for the given sequences and time range.")
        print("No file created.")
        return

    # --- Export ---
    if args.format == "csv":
        row_count = export_csv(rows, output_path)
    else:
        row_count = export_xlsx(rows, output_path)

    print(f"✔ Exported {row_count:,} rows to {output_path}")


if __name__ == "__main__":
    main()
